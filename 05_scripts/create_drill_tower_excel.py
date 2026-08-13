import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import pandas as pd
import os

def create_drill_tower_excel_report():
    print("=" * 80)
    print("Generating North Sea Drill Tower Excel Master Report...")
    print("=" * 80)

    base_dir = "C:/Users/frank/Desktop/EVM"
    raw_dir = os.path.join(base_dir, "01_raw_data")
    pbi_dir = os.path.join(base_dir, "03_power_bi")
    out_file = os.path.join(raw_dir, "Drill_Tower_EVM_Report.xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    total_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )

    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    red_font = Font(color="DC2626", bold=True)
    amber_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    amber_font = Font(color="D97706", bold=True)
    green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    green_font = Font(color="059669", bold=True)

    # -------------------------------------------------------------------------
    # TAB 1: EXECUTIVE DASHBOARD
    # -------------------------------------------------------------------------
    ws_dash = wb.create_sheet(title="Executive_Dashboard")
    ws_dash["A1"] = "North Sea Oil Platform Drill Tower Construction - Executive EVM Outturn Report"
    ws_dash["A1"].font = title_font

    ws_dash["A3"] = "Metric Name"
    ws_dash["B3"] = "Value ($ / %)"
    ws_dash["C3"] = "Status / Notes"
    ws_dash["A3"].font = header_font; ws_dash["A3"].fill = header_fill
    ws_dash["B3"].font = header_font; ws_dash["B3"].fill = header_fill
    ws_dash["C3"].font = header_font; ws_dash["C3"].fill = header_fill

    kpi_metrics = [
        ("Total Baseline Budget (BAC)", 26500000, "$#,##0", "Approved Scope Anchor"),
        ("Cumulative Planned Value (PV)", 20500000, "$#,##0", "BCWS Baseline Plan (Month 8)"),
        ("Cumulative Earned Value (EV)", 17540000, "$#,##0", "BCWP Physical Progress (60.8%)"),
        ("Cumulative Actual Cost (AC)", 23440000, "$#,##0", "ACWP Ledger Spend Incurred"),
        ("Cost Variance (CV = EV - AC)", -5900000, "$#,##0", "Critical Overrun"),
        ("Schedule Variance (SV = EV - PV)", -2960000, "$#,##0", "Schedule Slippage"),
        ("Cost Performance Index (CPI)", 0.7483, "0.00", "Critical Cost Efficiency Deficit"),
        ("Schedule Performance Index (SPI)", 0.8556, "0.00", "Schedule Execution Delay"),
        ("Critical Ratio (CR = CPI * SPI)", 0.6402, "0.00", "Critical Combined Deficit"),
        ("Outturn Forecast (EAC = BAC / CPI)", 35413604, "$#,##0", "Projected Final Outturn"),
        ("Variance at Completion (VAC = BAC - EAC)", -8913604, "$#,##0", "Projected Final Deficit"),
        ("Earned Schedule (ES)", 7.40, "0.00 Mos", "Lipke Time-Based Progress"),
        ("Time-Based Index SPI(t)", 0.9250, "0.00", "Time Variance: -18 Days")
    ]

    for idx, (m_name, val, fmt, note) in enumerate(kpi_metrics, start=4):
        ws_dash.cell(row=idx, column=1, value=m_name).font = bold_font
        c_val = ws_dash.cell(row=idx, column=2, value=val)
        c_val.font = bold_font; c_val.number_format = fmt
        c_note = ws_dash.cell(row=idx, column=3, value=note)
        c_note.font = regular_font
        if "Critical" in note or "Deficit" in note or "Overrun" in note or "Slippage" in note:
            c_val.fill = red_fill; c_val.font = red_font
        elif "Progress" in note or "Anchor" in note:
            c_val.fill = green_fill; c_val.font = green_font

    # Control Account Summary Table on Dashboard
    ws_dash["A18"] = "Control Account EVM Performance Matrix (Month 8 Status Date)"
    ws_dash["A18"].font = Font(name="Calibri", size=12, bold=True, color="1F4E79")

    df_wbs = pd.read_csv(os.path.join(pbi_dir, "Dim_WBS.csv"))
    wbs_headers = ["Task_ID", "WBS_Code", "Task_Name", "CAM", "BAC ($)"]
    for col_i, h in enumerate(wbs_headers, start=1):
        c = ws_dash.cell(row=20, column=col_i, value=h)
        c.font = header_font; c.fill = header_fill

    for r_i, r in df_wbs.iterrows():
        row_num = 21 + r_i
        ws_dash.cell(row=row_num, column=1, value=r["Task_ID"]).font = bold_font
        ws_dash.cell(row=row_num, column=2, value=r["WBS_Code"]).font = regular_font
        ws_dash.cell(row=row_num, column=3, value=r["Task_Name"]).font = regular_font
        ws_dash.cell(row=row_num, column=4, value=r["CAM"]).font = regular_font
        c_tbc = ws_dash.cell(row=row_num, column=5, value=r["TBC"])
        c_tbc.font = regular_font; c_tbc.number_format = "$#,##0"

    # Total Row
    tot_row = 21 + len(df_wbs)
    ws_dash.cell(row=tot_row, column=1, value="Total").font = bold_font
    ws_dash.cell(row=tot_row, column=3, value="Total Portfolio Scope").font = bold_font
    c_tot = ws_dash.cell(row=tot_row, column=5, value="=SUM(E21:E30)")
    c_tot.font = bold_font; c_tot.number_format = "$#,##0"; c_tot.border = total_border

    # -------------------------------------------------------------------------
    # TAB 2: PV BASELINE (01_PV_Baseline.csv)
    # -------------------------------------------------------------------------
    ws_pv = wb.create_sheet(title="PV_Baseline")
    df_pv = pd.read_csv(os.path.join(raw_dir, "01_PV_Baseline.csv"))
    ws_pv.append(list(df_pv.columns))
    for _, r in df_pv.iterrows():
        ws_pv.append(list(r))

    # -------------------------------------------------------------------------
    # TAB 3: EV PROGRESS (02_EV_Progress.csv)
    # -------------------------------------------------------------------------
    ws_ev = wb.create_sheet(title="EV_Progress")
    df_ev = pd.read_csv(os.path.join(raw_dir, "02_EV_Progress.csv"))
    ws_ev.append(list(df_ev.columns))
    for _, r in df_ev.iterrows():
        ws_ev.append(list(r))

    # Add openpyxl RAG Conditional Formatting on EV Progress % Complete
    ws_ev.conditional_formatting.add("E2:P11", CellIsRule(operator='equal', formula=['1.0'], fill=green_fill, font=green_font))
    ws_ev.conditional_formatting.add("E2:P11", CellIsRule(operator='between', formula=['0.8', '0.999'], fill=amber_fill, font=amber_font))
    ws_ev.conditional_formatting.add("E2:P11", CellIsRule(operator='lessThan', formula=['0.8'], fill=red_fill, font=red_font))

    # -------------------------------------------------------------------------
    # TAB 4: AC ACTUALS (03_AC_Actuals.csv)
    # -------------------------------------------------------------------------
    ws_ac = wb.create_sheet(title="AC_Actuals")
    df_ac = pd.read_csv(os.path.join(raw_dir, "03_AC_Actuals.csv"))
    ws_ac.append(list(df_ac.columns))
    for _, r in df_ac.iterrows():
        ws_ac.append(list(r))

    # -------------------------------------------------------------------------
    # TAB 5: GANTT SCHEDULE & DEPENDENCIES
    # -------------------------------------------------------------------------
    ws_gantt = wb.create_sheet(title="Gantt_Schedule")
    df_gantt = pd.read_csv(os.path.join(pbi_dir, "Fact_Gantt_Schedule.csv"))
    ws_gantt.append(list(df_gantt.columns))
    for _, r in df_gantt.iterrows():
        ws_gantt.append(list(r))

    # -------------------------------------------------------------------------
    # FORMATTING ALL SHEETS
    # -------------------------------------------------------------------------
    for sheet in wb.worksheets:
        if sheet.title == "Executive_Dashboard":
            continue
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center" if "Month" in str(cell.value) or "Code" in str(cell.value) else "left")
        
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for col_idx, cell in enumerate(row, start=1):
                cell.font = regular_font
                cell.border = thin_border
                header_val = sheet.cell(row=1, column=col_idx).value or ""
                if "TBC" in header_val or "AC" in header_val or ("Month_" in header_val and "%" not in header_val and sheet.title == "PV_Baseline"):
                    cell.number_format = "$#,##0"
                    cell.alignment = Alignment(horizontal="right")
                elif "%" in header_val:
                    cell.number_format = "0.0%"
                    cell.alignment = Alignment(horizontal="right")
                elif "ID" in header_val or "Code" in header_val or "Date" in header_val:
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")

        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    try:
        wb.save(out_file)
        print(f"Successfully generated Drill_Tower_EVM_Report.xlsx at: {out_file}")
    except PermissionError:
        alt_out = out_file.replace(".xlsx", "_updated.xlsx")
        wb.save(alt_out)
        print(f"[NOTE] Original file was locked; saved updated report to: {alt_out}")

if __name__ == "__main__":
    create_drill_tower_excel_report()
