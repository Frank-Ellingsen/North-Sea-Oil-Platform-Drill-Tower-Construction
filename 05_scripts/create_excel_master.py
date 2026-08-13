import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
bold_font = Font(name="Calibri", size=11, bold=True)
regular_font = Font(name="Calibri", size=11)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
total_top_border = Side(style='thin', color='000000')
total_bottom_border = Side(style='double', color='000000')
total_border = Border(top=total_top_border, bottom=total_bottom_border)

# 1. Planned Value Baseline Sheet
ws_pv = wb.create_sheet(title="PV_Baseline")
pv_headers = ["Item_ID", "WBS_Code", "Task_Name", "TBC", "Month_1", "Month_2", "Month_3", "Month_4", "Month_5", "Month_6"]
pv_data = [
    ["T101", "1.1.1", "Systems Engineering", 120000, 40000, 50000, 30000, 0, 0, 0],
    ["T102", "1.1.2", "Procurement", 300000, 100000, 150000, 50000, 0, 0, 0],
    ["T103", "1.2.1", "Site Excavation", 180000, 0, 20000, 80000, 60000, 20000, 0],
    ["T104", "1.2.2", "Foundation Concrete", 240000, 0, 0, 40000, 120000, 80000, 0],
    ["T105", "1.3.1", "Structural Assembly", 360000, 0, 0, 0, 60000, 200000, 100000]
]
ws_pv.append(pv_headers)
for row in pv_data:
    ws_pv.append(row)
ws_pv.append(["Total", "", "Total Baseline Budget", "=SUM(D2:D6)", "=SUM(E2:E6)", "=SUM(F2:F6)", "=SUM(G2:G6)", "=SUM(H2:H6)", "=SUM(I2:I6)", "=SUM(J2:J6)"])

# 2. Progress EV Sheet
ws_ev = wb.create_sheet(title="Progress_EV")
ev_headers = ["Item_ID", "WBS_Code", "Task_Name", "TBC", "Month_1_%", "Month_2_%", "Month_3_%", "Month_4_%", "Month_5_%", "Month_6_%"]
ev_data = [
    ["T101", "1.1.1", "Systems Engineering", 120000, 0.3333, 0.7500, 1.0000, 1.0000, 1.0000, 1.0000],
    ["T102", "1.1.2", "Procurement", 300000, 0.1000, 0.5000, 0.8000, 0.8000, 0.9000, 1.0000],
    ["T103", "1.2.1", "Site Excavation", 180000, 0.0000, 0.1111, 0.5556, 0.8889, 1.0000, 1.0000],
    ["T104", "1.2.2", "Foundation Concrete", 240000, 0.0000, 0.0000, 0.1000, 0.4000, 0.8000, 1.0000],
    ["T105", "1.3.1", "Structural Assembly", 360000, 0.0000, 0.0000, 0.0000, 0.1500, 0.6000, 0.9000]
]
ws_ev.append(ev_headers)
for row in ev_data:
    ws_ev.append(row)

# 3. Actual Costs Sheet
ws_ac = wb.create_sheet(title="Actual_Costs")
ac_headers = ["Item_ID", "WBS_Code", "Task_Name", "TBC", "Month_1_AC", "Month_2_AC", "Month_3_AC", "Month_4_AC", "Month_5_AC", "Month_6_AC"]
ac_data = [
    ["T101", "1.1.1", "Systems Engineering", 120000, 42000, 51000, 32000, 0, 0, 0],
    ["T102", "1.1.2", "Procurement", 300000, 95000, 160000, 55000, 5000, 12000, 28000],
    ["T103", "1.2.1", "Site Excavation", 180000, 0, 22000, 85000, 45000, 21000, 0],
    ["T104", "1.2.2", "Foundation Concrete", 240000, 0, 0, 38000, 92000, 84000, 41000],
    ["T105", "1.3.1", "Structural Assembly", 360000, 0, 0, 0, 55000, 185000, 96000]
]
ws_ac.append(ac_headers)
for row in ac_data:
    ws_ac.append(row)
ws_ac.append(["Total", "", "Total Actual Costs", "=SUM(D2:D6)", "=SUM(E2:E6)", "=SUM(F2:F6)", "=SUM(G2:G6)", "=SUM(H2:H6)", "=SUM(I2:I6)", "=SUM(J2:J6)"])

# 4. Dim_WBS Sheet
ws_wbs = wb.create_sheet(title="Dim_WBS")
wbs_headers = ["Task_ID", "WBS_Code", "WBS_Level_1", "WBS_Level_2", "Task_Name", "CAM", "TBC"]
wbs_data = [
    ["T101", "1.1.1", "1.0 Engineering", "1.1 Systems", "Systems Engineering", "L. Hansen", 120000],
    ["T102", "1.1.2", "1.0 Engineering", "1.1 Systems", "Procurement", "A. Olsen", 300000],
    ["T103", "1.2.1", "1.0 Construction", "1.2 Civil Works", "Site Excavation", "E. Johansen", 180000],
    ["T104", "1.2.2", "1.0 Construction", "1.2 Civil Works", "Foundation Concrete", "E. Johansen", 240000],
    ["T105", "1.3.1", "1.0 Construction", "1.3 Structural", "Structural Assembly", "K. Andersen", 360000]
]
ws_wbs.append(wbs_headers)
for row in wbs_data:
    ws_wbs.append(row)

# 5. Bastick Parametric S-Curve Model Sheet
ws_scurve = wb.create_sheet(title="Bastick_S_Curve_Model")
ws_scurve["A1"] = "Liam Bastick Parametric S-Curve Model"
ws_scurve["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F4E79")

assumptions = [
    ("Initial_Percentage_Completed", 0.00, "0.0%"),
    ("Target_Percentage", 1.00, "100.0%"),
    ("Exp_Factor", 2.50, "0.00"),
    ("Exp_Growth_Start_Month_No", 1, "0"),
    ("Exp_Duration_in_Months", 6, "0"),
    ("Total_Budget_Cost (BAC)", 1200000, "$#,##0")
]

ws_scurve["A3"] = "Parameter Name"
ws_scurve["B3"] = "Value"
ws_scurve["A3"].font = header_font
ws_scurve["A3"].fill = header_fill
ws_scurve["B3"].font = header_font
ws_scurve["B3"].fill = header_fill

for idx, (param, val, fmt) in enumerate(assumptions, start=4):
    ws_scurve.cell(row=idx, column=1, value=param).font = bold_font
    cell = ws_scurve.cell(row=idx, column=2, value=val)
    cell.font = regular_font
    cell.number_format = fmt

timeline_start_row = 12
ws_scurve.cell(row=timeline_start_row, column=1, value="Metric / Month").font = header_font
ws_scurve.cell(row=timeline_start_row, column=1).fill = header_fill

months = [1, 2, 3, 4, 5, 6]
for col_idx, m in enumerate(months, start=2):
    c = ws_scurve.cell(row=timeline_start_row, column=col_idx, value=f"Month {m}")
    c.font = header_font
    c.fill = header_fill

ws_scurve.cell(row=13, column=1, value="Month Index (t)").font = bold_font
for col_idx, m in enumerate(months, start=2):
    ws_scurve.cell(row=13, column=col_idx, value=m).font = regular_font

ws_scurve.cell(row=14, column=1, value="Cumulative % Complete (S-Curve)").font = bold_font
for col_idx, m in enumerate(months, start=2):
    col_let = get_column_letter(col_idx)
    formula = f"= $B$4 + ($B$5 - $B$4) / (1 + $B$6^(($B$7 + $B$8/2 - {col_let}13) / $B$8))"
    c = ws_scurve.cell(row=14, column=col_idx, value=formula)
    c.font = regular_font
    c.number_format = "0.0%"

ws_scurve.cell(row=15, column=1, value="Cumulative Planned Value ($)").font = bold_font
for col_idx, m in enumerate(months, start=2):
    col_let = get_column_letter(col_idx)
    formula = f"= {col_let}14 * $B$9"
    c = ws_scurve.cell(row=15, column=col_idx, value=formula)
    c.font = regular_font
    c.number_format = "$#,##0"

ws_scurve.cell(row=16, column=1, value="Incremental Planned Value ($)").font = bold_font
for col_idx, m in enumerate(months, start=2):
    col_let = get_column_letter(col_idx)
    if col_idx == 2:
        formula = f"= {col_let}15"
    else:
        prev_let = get_column_letter(col_idx - 1)
        formula = f"= {col_let}15 - {prev_let}15"
    c = ws_scurve.cell(row=16, column=col_idx, value=formula)
    c.font = regular_font
    c.number_format = "$#,##0"

for sheet in wb.worksheets:
    if sheet.title == "Bastick_S_Curve_Model":
        continue
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if "Month" in str(cell.value) or "Code" in str(cell.value) else "left")
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
        is_total = (row[0].value == "Total")
        for col_idx, cell in enumerate(row, start=1):
            if is_total:
                cell.font = bold_font
                cell.border = total_border
            else:
                cell.font = regular_font
                cell.border = thin_border
            
            header_val = sheet.cell(row=1, column=col_idx).value or ""
            if "TBC" in header_val or "AC" in header_val or ("Month_" in header_val and "%" not in header_val and sheet.title == "PV_Baseline"):
                cell.number_format = "$#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif "%" in header_val:
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="right")
            elif "ID" in header_val or "Code" in header_val:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in sheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

from openpyxl.formatting.rule import CellIsRule

# Apply RAG Conditional Formatting on Progress_EV Sheet (% Complete)
red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
red_font = Font(color="DC2626", bold=True)
amber_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
amber_font = Font(color="D97706", bold=True)
green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
green_font = Font(color="059669", bold=True)

# 100% Complete -> Green
ws_ev.conditional_formatting.add("E2:J6", CellIsRule(operator='equal', formula=['1.0'], fill=green_fill, font=green_font))
# 80% to 99% -> Amber
ws_ev.conditional_formatting.add("E2:J6", CellIsRule(operator='between', formula=['0.8', '0.999'], fill=amber_fill, font=amber_font))
# Less than 80% -> Red
ws_ev.conditional_formatting.add("E2:J6", CellIsRule(operator='lessThan', formula=['0.8'], fill=red_fill, font=red_font))

out_file = "C:/Users/frank/Desktop/EVM/01_raw_data/EVM_Master_Data.xlsx"
try:
    wb.save(out_file)
    print("Successfully generated EVM_Master_Data.xlsx with RAG formatting at:", out_file)
except PermissionError:
    alt_out = out_file.replace(".xlsx", "_updated.xlsx")
    wb.save(alt_out)
    print("[NOTE] Original file was locked; saved updated report to:", alt_out)
