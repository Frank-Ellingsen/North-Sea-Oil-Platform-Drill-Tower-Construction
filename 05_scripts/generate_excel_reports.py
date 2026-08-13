import os
import sys

# Ensure UTF-8 console output for Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def build_excel_report():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    docs_dir = os.path.join(base_dir, "06_docs")
    raw_dir = os.path.join(base_dir, "01_raw_data")
    os.makedirs(docs_dir, exist_ok=True)
    os.makedirs(raw_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles
    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="D1D5DB")
    font_sec_hdr = Font(name="Segoe UI", size=12, bold=True, color="1F2937")
    font_tbl_hdr = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Segoe UI", size=10, bold=True, color="111827")
    font_regular = Font(name="Segoe UI", size=10, color="1F2937")
    font_muted = Font(name="Segoe UI", size=9, color="6B7280")
    
    font_red = Font(name="Segoe UI", size=10, bold=True, color="DC2626")
    font_amber = Font(name="Segoe UI", size=10, bold=True, color="D97706")
    font_green = Font(name="Segoe UI", size=10, bold=True, color="059669")

    fill_header = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    fill_sec_hdr = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    fill_zebra = PatternFill(start_color="FAFAFA", end_color="FAFAFA", fill_type="solid")
    fill_red = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    fill_amber = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    fill_green = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="E5E7EB")
    thick_bottom_side = Side(border_style="medium", color="111827")
    double_bottom_side = Side(border_style="double", color="111827")

    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_header = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
    border_total = Border(top=thin_border_side, bottom=double_bottom_side)

    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_center = Alignment(horizontal="center", vertical="center")

    fmt_currency = "$#,##0"
    fmt_currency_exact = "$#,##0.00"
    fmt_pct = "0.0%"
    fmt_index = "0.000"
    fmt_int = "#,##0"

    def apply_header_banner(ws, title, subtitle):
        ws.merge_cells("A1:H1")
        ws.merge_cells("A2:H2")
        
        cell_t = ws["A1"]
        cell_t.value = f"  {title}"
        cell_t.font = font_title
        cell_t.fill = fill_header
        cell_t.alignment = align_left

        cell_s = ws["A2"]
        cell_s.value = f"  {subtitle}"
        cell_s.font = font_subtitle
        cell_s.fill = fill_header
        cell_s.alignment = align_left

        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 18

    # =========================================================================
    # TAB 1: EXECUTIVE BRIEFING
    # =========================================================================
    ws1 = wb.create_sheet(title="Executive Briefing")
    ws1.views.sheetView[0].showGridLines = True
    apply_header_banner(ws1, "North Sea Platform Drill Tower Construction", "Executive Steering Committee Monthly Briefing | Status Date: Aug 31, 2026")

    # Overall Health Callout Banner
    ws1.merge_cells("A4:H4")
    cell_h = ws1["A4"]
    cell_h.value = "🚨 OVERALL HEALTH: CRITICAL COST OVERRUN & SCHEDULE SLIPPAGE (RED)  |  BAC Budget Burn Out: Month 9 (Sep 2026)"
    cell_h.font = font_red
    cell_h.fill = fill_red
    cell_h.alignment = align_left
    ws1.row_dimensions[4].height = 24

    # EVM Key Metrics Table
    ws1.cell(row=6, column=1, value="1. Key EVM Performance Metrics").font = font_sec_hdr
    
    headers1 = ["Metric Name", "Acronym", "Current Value", "Baseline Target", "Variance ($ / %)", "Performance Index", "Status Health", "Context & Threshold Notes"]
    ws1.row_dimensions[7].height = 24
    for c_idx, h in enumerate(headers1, start=1):
        cell = ws1.cell(row=7, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_header
        cell.alignment = align_center if c_idx in [2,6,7] else (align_right if c_idx in [3,4,5] else align_left)
        cell.border = border_header

    metrics_data = [
        ["Approved Baseline Budget", "BAC", 26500000, 26500000, 0, 1.000, "Approved Target", "Original baseline budget ceiling"],
        ["Earned Work Complete", "EV", 17540000, 26500000, "=C9-D9", "=C9/D9", "66.2% Complete", "Earned physical work accomplished"],
        ["Planned Baseline Target", "PV", 20500000, 26500000, "=C10-D10", "=C10/D10", "77.4% Scheduled", "Scheduled baseline progress to date"],
        ["Cumulative Actual Spend", "AC", 23440000, 26500000, "=C11-C9", "=C9/C11", "Severe Overrun", "Actual ledger expenditure ($5.90M > EV)"],
        ["Cost Variance", "CV", "=$C$9-$C$11", 0, "=$C$12", "=C9/C11", "🚨 CRITICAL", "Overrun of -$5.90M (CPI = 0.748)"],
        ["Schedule Variance", "SV", "=$C$9-$C$10", 0, "=$C$13", "=C9/C10", "🟡 DELAYED", "Value delay of -$2.96M (SPI = 0.856)"],
        ["Earned Schedule Velocity", "SPIt", 0.925, 1.000, -18.2, 0.925, "🟡 -18 Days", "Time-based schedule velocity (ES 7.40M)"],
        ["To-Complete Index (BAC)", "TCPI_BAC", "=(C8-C9)/(C8-C11)", 1.000, 5.07, 6.07, "🚨 UNVIABLE", "Requires impossible 607% efficiency"]
    ]

    for r_offset, r_data in enumerate(metrics_data, start=8):
        ws1.row_dimensions[r_offset].height = 20
        for c_offset, val in enumerate(r_data, start=1):
            cell = ws1.cell(row=r_offset, column=c_offset, value=val)
            cell.font = font_regular
            cell.border = border_cell
            
            if c_offset in [1, 8]:
                cell.alignment = align_left
            elif c_offset in [2, 7]:
                cell.alignment = align_center
            else:
                cell.alignment = align_right

            if c_offset in [3, 4, 5] and isinstance(val, (int, float)):
                cell.number_format = fmt_currency
            elif c_offset == 6 and isinstance(val, (int, float)):
                cell.number_format = fmt_index
            elif str(val).startswith("="):
                if c_offset in [3, 4, 5]:
                    cell.number_format = fmt_currency
                elif c_offset == 6:
                    cell.number_format = fmt_index

            if c_offset == 7:
                if "CRITICAL" in str(val) or "Severe" in str(val) or "UNVIABLE" in str(val):
                    cell.fill = fill_red
                    cell.font = font_red
                elif "DELAYED" in str(val) or "18 Days" in str(val):
                    cell.fill = fill_amber
                    cell.font = font_amber
                elif "Complete" in str(val) or "Approved" in str(val):
                    cell.fill = fill_green
                    cell.font = font_green

    # Section 2: Outturn Predictions & P90 Risk
    ws1.cell(row=18, column=1, value="2. Outturn Predictions & P90 Risk Reserve").font = font_sec_hdr
    headers_scen = ["Scenario", "EAC Outturn ($)", "Variance at Completion (VAC)", "Completion Date", "Schedule Slip", "Risk Reserve Required ($)"]
    ws1.row_dimensions[19].height = 22
    for c_idx, h in enumerate(headers_scen, start=1):
        cell = ws1.cell(row=19, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_header
        cell.alignment = align_center if c_idx in [4,5] else (align_right if c_idx in [2,3,6] else align_left)
        cell.border = border_header

    scen_data = [
        ["Deterministic Base EAC", 35413604, "=26500000-B20", "2027-01-31", "+31 Days", 0],
        ["Monte Carlo P50 Forecast", 34060783, "=26500000-B21", "2027-02-27", "+57 Days", 0],
        ["Monte Carlo P90 Risk Ceiling", 35815202, "=26500000-B22", "2027-03-14", "+72 Days", "=B22-B20"]
    ]
    for r_offset, r_data in enumerate(scen_data, start=20):
        ws1.row_dimensions[r_offset].height = 20
        for c_offset, val in enumerate(r_data, start=1):
            cell = ws1.cell(row=r_offset, column=c_offset, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_offset in [2, 3, 6]:
                cell.alignment = align_right
                cell.number_format = fmt_currency
            elif c_offset in [4, 5]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left
            
            if r_offset == 22:
                cell.fill = fill_red
                cell.font = font_red

    # =========================================================================
    # TAB 2: RUNWAY & CASH BURN
    # =========================================================================
    ws2 = wb.create_sheet(title="Runway & Cash Burn")
    ws2.views.sheetView[0].showGridLines = True
    apply_header_banner(ws2, "Monthly Cash Burn Speed & Budget Depletion Forecast", "Exact BAC Budget Exhaustion Point: Month 9 (September 2026)")

    # Cards Summary Block
    cards_data = [
        ("Avg Monthly Cash Burn", 2930000, "Baseline Plan: $2.56M/Month (+14.5% faster)", fill_red, font_red),
        ("Remaining Baseline Capital", 3060000, "$23.44M spent of $26.50M BAC budget", fill_amber, font_amber),
        ("Budget Burn Out Month", "Month 9 (Sep 2026)", "100% of $BAC budget ceiling exhausted", fill_red, font_red),
        ("Required Overrun Financing", 8913604, "Needs approval before Sep 15, 2026", fill_red, font_red)
    ]
    for idx, (label, val, sub, fill_c, font_c) in enumerate(cards_data):
        col_start = 1 + (idx * 2)
        ws2.merge_cells(start_row=4, start_column=col_start, end_row=4, end_column=col_start+1)
        ws2.merge_cells(start_row=5, start_column=col_start, end_row=5, end_column=col_start+1)
        ws2.merge_cells(start_row=6, start_column=col_start, end_row=6, end_column=col_start+1)
        
        c_lbl = ws2.cell(row=4, column=col_start, value=label.upper())
        c_lbl.font = font_muted
        c_lbl.alignment = align_center

        c_val = ws2.cell(row=5, column=col_start, value=val)
        c_val.font = font_c
        c_val.alignment = align_center
        if isinstance(val, (int, float)):
            c_val.number_format = fmt_currency

        c_sub = ws2.cell(row=6, column=col_start, value=sub)
        c_sub.font = font_muted
        c_sub.alignment = align_center

        for r in range(4, 7):
            for c in range(col_start, col_start+2):
                ws2.cell(row=r, column=c).fill = fill_c
                ws2.cell(row=r, column=c).border = border_cell

    # Monthly Burn Table
    ws2.cell(row=8, column=1, value="Monthly Cash Burn Speed ($PV, $EV, $AC) & Budget Exhaustion Schedule").font = font_sec_hdr
    b_headers = ["Month #", "Calendar Month", "Planned Monthly ($PV)", "Earned Monthly ($EV)", "Actual Monthly ($AC)", "Cumulative Spend ($AC)", "Remaining Capital ($BAC - AC)", "Runway Status & Depletion Note"]
    ws2.row_dimensions[9].height = 24
    for c_idx, h in enumerate(b_headers, start=1):
        cell = ws2.cell(row=9, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_header
        cell.alignment = align_center if c_idx in [1,2] else (align_right if c_idx in [3,4,5,6,7] else align_left)
        cell.border = border_header

    monthly_burns = [
        [1, "Month 1 (Jan 2026)", 300000, 300000, 320000, "=E10", "=26500000-F10", "Detail steel engineering ramp-up"],
        [2, "Month 2 (Feb 2026)", 900000, 840000, 950000, "=F10+E11", "=26500000-F11", "Engineering completed on baseline"],
        [3, "Month 3 (Mar 2026)", 2700000, 2260000, 2900000, "=F11+E12", "=26500000-F12", "Procurement placement delays"],
        [4, "Month 4 (Apr 2026)", 3800000, 3520000, 4050000, "=F12+E13", "=26500000-F13", "Verdal yard sub-structure start"],
        [5, "Month 5 (May 2026)", 3000000, 2900000, 3220000, "=F13+E14", "=26500000-F14", "Sub-structure assembly"],
        [6, "Month 6 (Jun 2026)", 4000000, 2750000, 4450000, "=F14+E15", "=26500000-F15", "Egersund mast fitting misalignment rework"],
        [7, "Month 7 (Jul 2026)", 3000000, 2620000, 3850000, "=F15+E16", "=26500000-F16", "Double-shift NDT welding labor"],
        [8, "Month 8 (Aug 2026)", 2800000, 2350000, 3700000, "=F16+E17", "=26500000-F17", "Status Date: $23.44M cumulative actual spend"],
        [9, "Month 9 (Sep 2026)", 2500000, 2100000, 3200000, "=F17+E18", "=26500000-F18", "💥 100% BAC Budget Exhausted (Burn Out)"],
        [10, "Month 10 (Oct 2026)", 1500000, 2200000, 3300000, "=F18+E19", "=26500000-F19", "Overrun Financing Phase (Heerema Marine)"]
    ]

    for r_offset, r_data in enumerate(monthly_burns, start=10):
        ws2.row_dimensions[r_offset].height = 20
        for c_offset, val in enumerate(r_data, start=1):
            cell = ws2.cell(row=r_offset, column=c_offset, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_offset in [1, 2]:
                cell.alignment = align_center
            elif c_offset in [3, 4, 5, 6, 7]:
                cell.alignment = align_right
                cell.number_format = fmt_currency
            else:
                cell.alignment = align_left
            
            if r_offset == 18:
                cell.fill = fill_red
                cell.font = font_red

    # =========================================================================
    # TAB 3: COST VARIANCE WATERFALL BRIDGE
    # =========================================================================
    ws3 = wb.create_sheet(title="Cost Variance Waterfall")
    ws3.views.sheetView[0].showGridLines = True
    apply_header_banner(ws3, "Cost Variance Waterfall Bridge ($BAC → EAC Outturn Breakdown)", "Step-by-Step Cost Variance Drivers from Baseline to Outturn Forecast")

    ws3.cell(row=4, column=1, value="Step-by-Step Variance Bridge per WBS Control Account").font = font_sec_hdr
    wf_headers = ["Step #", "WBS Control Account / Driver", "Step Variance ($)", "Cumulative Total ($)", "Variance Category", "Root Cause Driver Explanation"]
    ws3.row_dimensions[5].height = 24
    for c_idx, h in enumerate(wf_headers, start=1):
        cell = ws3.cell(row=5, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_header
        cell.alignment = align_center if c_idx in [1,5] else (align_right if c_idx in [3,4] else align_left)
        cell.border = border_header

    wf_steps = [
        [0, "Approved Baseline Budget ($BAC)", 26500000, "=C6", "Baseline Target", "Approved original contract baseline budget"],
        [1, "WBS 1.1 Detail Engineering", 300000, "=D6+C7", "Scope Rework", "Minor drafting revision hours"],
        [2, "WBS 1.2 Procurement", 600000, "=D7+C8", "Material Inflation", "High-grade tubular steel market price surge"],
        [3, "WBS 1.3.1 Verdal Yard Fabrication", 200000, "=D8+C9", "Yard Overtime", "Welder shift overtime"],
        [4, "WBS 1.3.2 Egersund Mast Assembly", 2400000, "=D9+C10", "Primary Rework Driver", "Pipe fitting misalignment & 24/7 NDT welding"],
        [5, "WBS 1.4.1 Heavy Lift Vessel Mobilization", 1300000, "=D10+C11", "Marine Standby", "Autumn sea-state vessel standby daily rates"],
        [6, "WBS 1.4.2 Topside Mating", 100000, "=D11+C12", "Offshore Fit", "Offshore trial fitting hours"],
        [7, "WBS 1.5 Commissioning & Hook-up", 200000, "=D12+C13", "System Test", "Pre-commissioning loops"],
        [8, "Time Delay Overhead Spread (SPIt)", 3813604, "=D13+C14", "Schedule Overhead", "Extending execution past Dec 31 spreads PMO & site overhead"],
        [9, "Final Outturn Forecast ($EAC)", 0, "=D14", "Total EAC Outturn", "Total Outturn Deficit: +$8.91M VAC (-33.6%)"]
    ]

    for r_offset, r_data in enumerate(wf_steps, start=6):
        ws3.row_dimensions[r_offset].height = 20
        for c_offset, val in enumerate(r_data, start=1):
            cell = ws3.cell(row=r_offset, column=c_offset, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_offset == 1:
                cell.alignment = align_center
            elif c_offset in [3, 4]:
                cell.alignment = align_right
                cell.number_format = fmt_currency
            elif c_offset == 5:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

            if r_offset in [10, 14]:
                cell.font = font_red
                if r_offset == 10:
                    cell.fill = fill_red
            if r_offset == 15:
                cell.font = font_bold
                cell.fill = fill_sec_hdr

    # =========================================================================
    # TAB 4: EVM & S-CURVE PERFORMANCE
    # =========================================================================
    ws4 = wb.create_sheet(title="EVM & S-Curve Performance")
    ws4.views.sheetView[0].showGridLines = True
    apply_header_banner(ws4, "Performance Measurement Baseline (S-Curves) & Cumulative EVM", "Direct Endpoint Metrics, Cost Variance & Time Slippage Analysis")

    ws4.cell(row=4, column=1, value="Monthly Cumulative Performance Measurement Baseline (PMB S-Curve)").font = font_sec_hdr
    pmb_headers = ["Month #", "Month Name", "Planned Value ($PV)", "Earned Value ($EV)", "Actual Cost ($AC)", "Cost Variance ($CV)", "Schedule Variance ($SV)", "CPI Index", "SPI Index"]
    ws4.row_dimensions[5].height = 24
    for c_idx, h in enumerate(pmb_headers, start=1):
        cell = ws4.cell(row=5, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_header
        cell.alignment = align_center if c_idx in [1,2] else (align_right if c_idx in [3,4,5,6,7,8,9] else align_left)
        cell.border = border_header

    pmb_data = [
        [1, "Jan 2026", 300000, 300000, 320000, "=D6-E6", "=D6-C6", "=D6/E6", "=D6/C6"],
        [2, "Feb 2026", 1200000, 1140000, 1270000, "=D7-E7", "=D7-C7", "=D7/E7", "=D7/C7"],
        [3, "Mar 2026", 3900000, 3400000, 4170000, "=D8-E8", "=D8-C8", "=D8/E8", "=D8/C8"],
        [4, "Apr 2026", 7700000, 6920000, 8220000, "=D9-E9", "=D9-C9", "=D9/E9", "=D9/C9"],
        [5, "May 2026", 10700000, 9820000, 11440000, "=D10-E10", "=D10-C10", "=D10/E10", "=D10/C10"],
        [6, "Jun 2026", 14700000, 12570000, 15890000, "=D11-E11", "=D11-C11", "=D11/E11", "=D11/C11"],
        [7, "Jul 2026", 17700000, 15190000, 19740000, "=D12-E12", "=D12-C12", "=D12/E12", "=D12/C12"],
        [8, "Aug 2026 (Status)", 20500000, 17540000, 23440000, "=D13-E13", "=D13-C13", "=D13/E13", "=D13/C13"]
    ]

    for r_offset, r_data in enumerate(pmb_data, start=6):
        ws4.row_dimensions[r_offset].height = 20
        for c_offset, val in enumerate(r_data, start=1):
            cell = ws4.cell(row=r_offset, column=c_offset, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_offset in [1, 2]:
                cell.alignment = align_center
            elif c_offset in [3, 4, 5, 6, 7]:
                cell.alignment = align_right
                cell.number_format = fmt_currency
            elif c_offset in [8, 9]:
                cell.alignment = align_right
                cell.number_format = fmt_index
            
            if r_offset == 13:
                if c_offset in [6, 8]:
                    cell.fill = fill_red
                    cell.font = font_red
                elif c_offset in [7, 9]:
                    cell.fill = fill_amber
                    cell.font = font_amber

    # =========================================================================
    # TAB 5: CFO FINANCIALS
    # =========================================================================
    ws5 = wb.create_sheet(title="CFO Financials")
    ws5.views.sheetView[0].showGridLines = True
    apply_header_banner(ws5, "CFO Control Account Financials & Outturn Forecast", "Control Account Level EAC Outturns, Cost Variances & Remaining ETC Liquidity")

    ws5.cell(row=4, column=1, value="Control Account Financial Performance & Liquidity Target").font = font_sec_hdr
    cfo_headers = ["WBS Code", "Control Account Name", "Baseline Budget ($BAC)", "Earned Value ($EV)", "Actual Cost ($AC)", "Cost Variance ($CV)", "CPI Index", "Outturn Forecast ($EAC)", "Outturn Deficit ($VAC)", "Remaining Liquidity ($ETC)"]
    ws5.row_dimensions[5].height = 24
    for c_idx, h in enumerate(cfo_headers, start=1):
        cell = ws5.cell(row=5, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_header
        cell.alignment = align_center if c_idx in [1,7] else (align_right if c_idx in [3,4,5,6,8,9,10] else align_left)
        cell.border = border_header

    cfo_data = [
        ["1.1.1", "Structural Steel Detail Engineering", 1200000, 1200000, 1290000, "=D6-E6", "=D6/E6", 1290000, "=C6-H6", "=H6-E6"],
        ["1.2.1", "High-Grade Tubular Steel Procurement", 3500000, 3500000, 3770000, "=D7-E7", "=D7/E7", 3770000, "=C7-H7", "=H7-E7"],
        ["1.3.1", "Yard Sub-Structure Fabrication (Verdal)", 4000000, 3600000, 4450000, "=D8-E8", "=D8/E8", 4944000, "=C8-H8", "=H8-E8"],
        ["1.3.2", "Derrick Tower Mast Assembly (Egersund)", 4800000, 3120000, 5450000, "=D9-E9", "=D9/E9", 8385000, "=C9-H9", "=H9-E9"]
    ]

    for r_offset, r_data in enumerate(cfo_data, start=6):
        ws5.row_dimensions[r_offset].height = 20
        for c_offset, val in enumerate(r_data, start=1):
            cell = ws5.cell(row=r_offset, column=c_offset, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_offset == 1:
                cell.alignment = align_center
            elif c_offset in [3, 4, 5, 6, 8, 9, 10]:
                cell.alignment = align_right
                cell.number_format = fmt_currency
            elif c_offset == 7:
                cell.alignment = align_right
                cell.number_format = fmt_index
            else:
                cell.alignment = align_left

            if r_offset == 9 and c_offset in [6, 9]:
                cell.fill = fill_red
                cell.font = font_red

    # =========================================================================
    # TAB 6: GANTT & CRITICAL PATH
    # =========================================================================
    ws6 = wb.create_sheet(title="Gantt & Critical Path")
    ws6.views.sheetView[0].showGridLines = True
    apply_header_banner(ws6, "Offshore EPC Gantt Schedule & Critical Path Logic", "10 Major Schedule Tasks, Predecessor Logic, Critical Path (8 Tasks) & Key Milestones")

    ws6.cell(row=4, column=1, value="EPC Schedule Master Activity & Critical Path Register").font = font_sec_hdr
    gantt_headers = ["Task ID", "WBS Code", "Task Deliverable / Milestone Name", "Predecessor", "Type", "Total Float", "% Complete", "Baseline Window", "Forecast Window", "Critical Path Flag"]
    ws6.row_dimensions[5].height = 24
    for c_idx, h in enumerate(gantt_headers, start=1):
        cell = ws6.cell(row=5, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_header
        cell.alignment = align_center if c_idx in [1,2,4,5,6,7,8,9,10] else align_left
        cell.border = border_header

    gantt_data = [
        ["T101", "1.1.1", "Structural Steel Detail Engineering", "-", "-", "0 Days", 1.00, "Jan 05 - Feb 28", "Jan 05 - Mar 15", "Yes"],
        ["M1", "1.1.0", "◆ Milestone: Engineering AFC Gate Review", "T101", "FS", "0 Days", 1.00, "Feb 28", "Achieved Mar 15", "Achieved Gate"],
        ["T102", "1.1.2", "Piping & Drilling Package Design", "T101", "FS", "+14 Days", 1.00, "Feb 01 - Mar 31", "Feb 01 - Apr 15", "No"],
        ["T103", "1.2.1", "High-Grade Tubular Steel Procurement", "T101 (+5D)", "FS", "0 Days", 1.00, "Mar 01 - Apr 30", "Mar 15 - May 31", "Yes (+31D Delay)"],
        ["T104", "1.2.2", "Mud Pumps & Top Drive Equipment", "T102", "FS", "+15 Days", 0.85, "Mar 15 - Jun 30", "Apr 01 - Jul 31", "No"],
        ["T105", "1.3.1", "Yard Sub-Structure Fabrication (Verdal)", "T103", "FS", "0 Days", 0.90, "May 01 - Jul 31", "Jun 01 - Aug 31", "Yes"],
        ["M3", "1.3.0", "◆ Milestone: Verdal Sub-Structure Handover", "T105", "FS", "0 Days", 0.90, "Jul 31", "Aug 31 Handover", "Active Gate"],
        ["T106", "1.3.2", "Derrick Tower Mast Assembly (Egersund)", "T105", "FS", "0 Days", 0.65, "Jun 01 - Aug 31", "Jul 01 - Sep 30", "Yes (-30D Float)"],
        ["M4", "1.3.3", "◆ Milestone: Derrick Mast Ready for Heavy Lift", "T106", "FS", "0 Days", 0.65, "Aug 31", "Sep 30 Target", "Critical Gate"],
        ["T107", "1.4.1", "Heavy Lift Vessel Mobilization (Heerema)", "T106", "FS", "0 Days", 0.30, "Aug 15 - Sep 15", "Sep 15 - Oct 15", "Yes"]
    ]

    for r_offset, r_data in enumerate(gantt_data, start=6):
        ws6.row_dimensions[r_offset].height = 20
        for c_offset, val in enumerate(r_data, start=1):
            cell = ws6.cell(row=r_offset, column=c_offset, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_offset in [1, 2, 4, 5, 6, 8, 9, 10]:
                cell.alignment = align_center
            elif c_offset == 7:
                cell.alignment = align_right
                cell.number_format = fmt_pct
            else:
                cell.alignment = align_left

            if c_offset == 10 and "Yes" in str(val):
                cell.fill = fill_red
                cell.font = font_red

    # Auto-adjust column widths for all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Skip merged banner row 1 & 2 in calculations
                if cell.row in [1, 2] or cell.value is None:
                    continue
                val_str = str(cell.value)
                if len(val_str) > max_len:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Save to both 06_docs and 01_raw_data
    file1 = os.path.join(docs_dir, "Drill_Tower_EVM_Master_Report.xlsx")
    file2 = os.path.join(raw_dir, "Drill_Tower_EVM_Master_Report.xlsx")
    wb.save(file1)
    wb.save(file2)
    print(f"✅ Generated: 06_docs/Drill_Tower_EVM_Master_Report.xlsx")
    print(f"✅ Generated: 01_raw_data/Drill_Tower_EVM_Master_Report.xlsx")

if __name__ == "__main__":
    build_excel_report()
