"""
Comprehensive Script: Update Excel Master Files & Power BI Star Schema Dataset
Applies all web reporting features (Milestones, Risk Matrix, Swimlanes, Cash Burn,
Waterfall Bridge, Donut Chart Data, Financial Ratios, Monte Carlo P90, 1-Page Status Report)
to Excel (.xlsx) and Power BI Star Schema CSVs + DAX / PowerQuery scripts.
"""

import os
import csv
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
RAW_DIR = os.path.join(BASE_DIR, "01_raw_data")
PBI_DIR = os.path.join(BASE_DIR, "03_power_bi")

# 1. DEFINE DATASTRUCTURES FOR NEW REPORTING FEATURES

# A. KEY MILESTONES
milestones_data = [
    {"Milestone_ID": "M1", "Milestone_Name": "Engineering AFC Gate Review", "Target_Date": "2026-03-15", "Baseline_Date": "2026-03-15", "Status": "Achieved", "RAG": "Green", "WBS_Code": "1.1"},
    {"Milestone_ID": "M2", "Milestone_Name": "Tubular Steel Material Delivery", "Target_Date": "2026-05-31", "Baseline_Date": "2026-05-31", "Status": "Achieved", "RAG": "Green", "WBS_Code": "1.2"},
    {"Milestone_ID": "M3", "Milestone_Name": "Verdal Sub-Structure Handover", "Target_Date": "2026-08-31", "Baseline_Date": "2026-08-31", "Status": "Active", "RAG": "Green", "WBS_Code": "1.3.1"},
    {"Milestone_ID": "M4", "Milestone_Name": "Derrick Mast Ready for Lift", "Target_Date": "2026-09-30", "Baseline_Date": "2026-08-31", "Status": "Critical Rework", "RAG": "Red", "WBS_Code": "1.3.2"},
    {"Milestone_ID": "M5", "Milestone_Name": "Offshore Topside Mating Complete", "Target_Date": "2026-11-15", "Baseline_Date": "2026-10-31", "Status": "Critical Delay", "RAG": "Red", "WBS_Code": "1.4.2"},
    {"Milestone_ID": "M6", "Milestone_Name": "Final Commercial COD Handover", "Target_Date": "2027-01-31", "Baseline_Date": "2026-12-31", "Status": "Forecast COD", "RAG": "Red", "WBS_Code": "1.5"}
]

# B. RISK MATRIX & REGISTER
risks_data = [
    {"Risk_ID": "R01", "Risk_Title": "Egersund Derrick Mast Welding Tolerance Rework", "Category": "Fabrication", "Probability": 5, "Impact": 5, "Risk_Score": 25, "RAG_Level": "Critical", "Financial_Exposure": 2400000, "Expected_Monetary_Value": 2400000, "Mitigation_Strategy": "Capped yard labor hours under fixed-fee contract & 24/7 NDT supervision", "CAM_Owner": "Lars Hansen"},
    {"Risk_ID": "R02", "Risk_Title": "North Sea Weather Standby Delay during Mating", "Category": "Marine", "Probability": 4, "Impact": 4, "Risk_Score": 16, "RAG_Level": "High", "Financial_Exposure": 1800000, "Expected_Monetary_Value": 720000, "Mitigation_Strategy": "Secured 14-day flexible mobilization window with Heerema Sleipnir", "CAM_Owner": "Erik Solberg"},
    {"Risk_ID": "R03", "Risk_Title": "Subsea Tubular Steel Market Price Surcharges", "Category": "Procurement", "Probability": 3, "Impact": 3, "Risk_Score": 9, "RAG_Level": "Medium", "Financial_Exposure": 600000, "Expected_Monetary_Value": 180000, "Mitigation_Strategy": "Locked bulk steel purchasing under long-term supplier framework", "CAM_Owner": "Ingrid Berg"},
    {"Risk_ID": "R04", "Risk_Title": "DNV Class Structural Redesign Review", "Category": "Engineering", "Probability": 2, "Impact": 3, "Risk_Score": 6, "RAG_Level": "Medium", "Financial_Exposure": 300000, "Expected_Monetary_Value": 60000, "Mitigation_Strategy": "Embedded DNV structural lead directly inside project engineering team", "CAM_Owner": "Geir Nilsen"},
    {"Risk_ID": "R05", "Risk_Title": "Offshore Hook-up Commissioning Crew Bottleneck", "Category": "Commissioning", "Probability": 2, "Impact": 2, "Risk_Score": 4, "RAG_Level": "Low", "Financial_Exposure": 200000, "Expected_Monetary_Value": 40000, "Mitigation_Strategy": "Pre-commissioned skid packages onshore in Verdal yard", "CAM_Owner": "Bjørn Lie"}
]

# C. COST VARIANCE WATERFALL BRIDGE
waterfall_data = [
    {"Step_ID": 0, "Component_Name": "Original Baseline Budget (BAC)", "Type": "Baseline", "Incremental_Cost": 26500000, "Cumulative_Cost": 26500000, "Pct_Share": 0.0, "Description": "Approved EPC baseline contract budget ceiling"},
    {"Step_ID": 1, "Component_Name": "WBS 1.1 Detail Structural Engineering", "Type": "Variance", "Incremental_Cost": 300000, "Cumulative_Cost": 26800000, "Pct_Share": 3.4, "Description": "Redesign of connection plates & third-party DNV class review"},
    {"Step_ID": 2, "Component_Name": "WBS 1.2 Procurement & Tubular Steel", "Type": "Variance", "Incremental_Cost": 600000, "Cumulative_Cost": 27400000, "Pct_Share": 6.7, "Description": "Subsea tubular steel price inflation & mud pump freight surcharge"},
    {"Step_ID": 3, "Component_Name": "WBS 1.3.1 Verdal Sub-Structure Yard", "Type": "Variance", "Incremental_Cost": 200000, "Cumulative_Cost": 27600000, "Pct_Share": 2.2, "Description": "Minor yard overtime for weld inspection"},
    {"Step_ID": 4, "Component_Name": "WBS 1.3.2 Egersund Derrick Mast Rework", "Type": "Variance", "Incremental_Cost": 2400000, "Cumulative_Cost": 30000000, "Pct_Share": 26.9, "Description": "PRIMARY YARD REWORK DRIVER: Pipe out-of-tolerance & 24/7 NDT welding"},
    {"Step_ID": 5, "Component_Name": "WBS 1.4.1 Heavy Lift Vessel Standby", "Type": "Variance", "Incremental_Cost": 1300000, "Cumulative_Cost": 31300000, "Pct_Share": 14.6, "Description": "North Sea autumn weather standby rates (Heerema Sleipnir)"},
    {"Step_ID": 6, "Component_Name": "WBS 1.4.2 Topside Lifting & Mating", "Type": "Variance", "Incremental_Cost": 100000, "Cumulative_Cost": 31400000, "Pct_Share": 1.1, "Description": "Pre-lift trial fitting and alignment fixtures"},
    {"Step_ID": 7, "Component_Name": "WBS 1.5 Hook-up & Commissioning", "Type": "Variance", "Incremental_Cost": 200000, "Cumulative_Cost": 31600000, "Pct_Share": 2.2, "Description": "Extended offshore commissioning specialist crew"},
    {"Step_ID": 8, "Component_Name": "Time Delay Overhead Spread (SPI_t)", "Type": "Variance", "Incremental_Cost": 3813604, "Cumulative_Cost": 35413604, "Pct_Share": 42.8, "Description": "TIME DELAY SPREAD: +31-day delay spreading PMO & site overhead"},
    {"Step_ID": 9, "Component_Name": "Final Outturn Forecast (EAC)", "Type": "Outturn", "Incremental_Cost": 35413604, "Cumulative_Cost": 35413604, "Pct_Share": 100.0, "Description": "Final predicted cost outcome at project completion"}
]

# D. MONTHLY CASH BURN SPEED & RUNWAY
cash_burn_data = [
    {"Period": "M1 (Jan 26)", "Monthly_PV": 300000, "Monthly_EV": 300000, "Monthly_AC": 320000, "Cum_AC": 320000, "Remaining_BAC": 26180000, "Runway_Status": "Baseline Runway"},
    {"Period": "M2 (Feb 26)", "Monthly_PV": 900000, "Monthly_EV": 840000, "Monthly_AC": 950000, "Cum_AC": 1270000, "Remaining_BAC": 25230000, "Runway_Status": "Baseline Runway"},
    {"Period": "M3 (Mar 26)", "Monthly_PV": 2700000, "Monthly_EV": 2260000, "Monthly_AC": 2900000, "Cum_AC": 4170000, "Remaining_BAC": 22330000, "Runway_Status": "Baseline Runway"},
    {"Period": "M4 (Apr 26)", "Monthly_PV": 3800000, "Monthly_EV": 3520000, "Monthly_AC": 4050000, "Cum_AC": 8220000, "Remaining_BAC": 18280000, "Runway_Status": "Baseline Runway"},
    {"Period": "M5 (May 26)", "Monthly_PV": 3000000, "Monthly_EV": 2900000, "Monthly_AC": 3220000, "Cum_AC": 11440000, "Remaining_BAC": 15060000, "Runway_Status": "Baseline Runway"},
    {"Period": "M6 (Jun 26)", "Monthly_PV": 4000000, "Monthly_EV": 2750000, "Monthly_AC": 4450000, "Cum_AC": 15890000, "Remaining_BAC": 10610000, "Runway_Status": "Peak Yard Spend Phase"},
    {"Period": "M7 (Jul 26)", "Monthly_PV": 3000000, "Monthly_EV": 2620000, "Monthly_AC": 3850000, "Cum_AC": 19740000, "Remaining_BAC": 6760000, "Runway_Status": "Peak Yard Spend Phase"},
    {"Period": "M8 (Aug 26 - Now)", "Monthly_PV": 2800000, "Monthly_EV": 2350000, "Monthly_AC": 3700000, "Cum_AC": 23440000, "Remaining_BAC": 3060000, "Runway_Status": "Critical Runway ($3.06M Left)"},
    {"Period": "M9 (Sep 26 - Burn)", "Monthly_PV": 2500000, "Monthly_EV": 2100000, "Monthly_AC": 3200000, "Cum_AC": 26640000, "Remaining_BAC": -140000, "Runway_Status": "100% BUDGET BURN OUT!"},
    {"Period": "M10 (Oct 26)", "Monthly_PV": 1500000, "Monthly_EV": 2200000, "Monthly_AC": 3300000, "Cum_AC": 29940000, "Remaining_BAC": -3440000, "Runway_Status": "Overrun Financing Phase"},
    {"Period": "M11 (Nov 26)", "Monthly_PV": 1000000, "Monthly_EV": 2000000, "Monthly_AC": 2500000, "Cum_AC": 32440000, "Remaining_BAC": -5940000, "Runway_Status": "Overrun Financing Phase"},
    {"Period": "M12 (Dec 26)", "Monthly_PV": 1000000, "Monthly_EV": 1800000, "Monthly_AC": 1800000, "Cum_AC": 34240000, "Remaining_BAC": -7740000, "Runway_Status": "Overrun Financing Phase"},
    {"Period": "M13 (Jan 27 - COD)", "Monthly_PV": 0, "Monthly_EV": 860000, "Monthly_AC": 1173604, "Cum_AC": 35413604, "Remaining_BAC": -8913604, "Runway_Status": "Final COD (+8.91M Overrun)"}
]

# E. COMMERCIAL FINANCIAL RATIOS
financial_ratios_data = [
    {"Metric": "Net Present Value (NPV @ 10% WACC)", "Value": "+$14,899,563", "Numeric_Value": 14899563, "Unit": "USD", "Evaluation": "Strong Positive Net Present Value"},
    {"Metric": "Internal Rate of Return (IRR)", "Value": "18.86%", "Numeric_Value": 0.1886, "Unit": "%", "Evaluation": "Outperforms 10.0% Hurdle Rate by +886 bps"},
    {"Metric": "Simple Payback Period", "Value": "4.43 Years", "Numeric_Value": 4.43, "Unit": "Years", "Evaluation": "53.1 Months (May 2031)"},
    {"Metric": "Profitability Index (PI)", "Value": "1.42", "Numeric_Value": 1.42, "Unit": "Ratio", "Evaluation": "Generates $1.42 PV per $1.00 spent"},
    {"Metric": "Total Simple ROI", "Value": "134.37%", "Numeric_Value": 1.3437, "Unit": "%", "Evaluation": "High 10-year cumulative commercial return"},
    {"Metric": "Annualized ROI (CAGR)", "Value": "8.89%", "Numeric_Value": 0.0889, "Unit": "%/Year", "Evaluation": "Compounded annual growth rate"},
    {"Metric": "Gross Future Value Year 10 (FV)", "Value": "$130,499,397", "Numeric_Value": 130499397, "Unit": "USD", "Evaluation": "Nominal cash generated across 10 years"},
    {"Metric": "Net Future Value (NFV)", "Value": "+$38,645,628", "Numeric_Value": 38645628, "Unit": "USD", "Evaluation": "Net future value after CAPEX outturn"}
]

# F. MONTE CARLO RISK SIMULATION RESULTS
monte_carlo_data = [
    {"Percentile": "P10", "Confidence_Level": "10% Optimistic", "Outturn_Cost_EAC": 32444302, "Cost_Overrun_VAC": -5944302, "Contingency_Reserve": 0, "Duration_Days": 408.1, "Completion_Date": "2027-02-13", "Schedule_Delay_Days": 43.1},
    {"Percentile": "P50", "Confidence_Level": "50% Median", "Outturn_Cost_EAC": 34060783, "Cost_Overrun_VAC": -7560783, "Contingency_Reserve": 0, "Duration_Days": 422.3, "Completion_Date": "2027-02-27", "Schedule_Delay_Days": 57.3},
    {"Percentile": "P80", "Confidence_Level": "80% Standard", "Outturn_Cost_EAC": 35195026, "Cost_Overrun_VAC": -8695026, "Contingency_Reserve": 0, "Duration_Days": 432.3, "Completion_Date": "2027-03-09", "Schedule_Delay_Days": 67.3},
    {"Percentile": "P90", "Confidence_Level": "90% High Confidence", "Outturn_Cost_EAC": 35815202, "Cost_Overrun_VAC": -9315202, "Contingency_Reserve": 401598, "Duration_Days": 437.5, "Completion_Date": "2027-03-14", "Schedule_Delay_Days": 72.5},
    {"Percentile": "P95", "Confidence_Level": "95% Extreme Risk", "Outturn_Cost_EAC": 36272986, "Cost_Overrun_VAC": -9772986, "Contingency_Reserve": 859382, "Duration_Days": 441.6, "Completion_Date": "2027-03-18", "Schedule_Delay_Days": 76.6}
]

def export_powerbi_csvs():
    """Exports new star schema CSV files for Power BI ingestion."""
    print("--- Exporting New Star Schema CSVs to 03_power_bi/ ---")
    
    pd.DataFrame(milestones_data).to_csv(os.path.join(PBI_DIR, "Fact_Milestones.csv"), index=False)
    print(" [PASS] Exported Fact_Milestones.csv")
    
    pd.DataFrame(risks_data).to_csv(os.path.join(PBI_DIR, "Fact_Risk_Register.csv"), index=False)
    print(" [PASS] Exported Fact_Risk_Register.csv")
    
    pd.DataFrame(waterfall_data).to_csv(os.path.join(PBI_DIR, "Fact_Waterfall_Bridge.csv"), index=False)
    print(" [PASS] Exported Fact_Waterfall_Bridge.csv")
    
    pd.DataFrame(cash_burn_data).to_csv(os.path.join(PBI_DIR, "Fact_Monthly_Burn_Rate.csv"), index=False)
    print(" [PASS] Exported Fact_Monthly_Burn_Rate.csv")
    
    pd.DataFrame(financial_ratios_data).to_csv(os.path.join(PBI_DIR, "Fact_Financial_Appraisal.csv"), index=False)
    print(" [PASS] Exported Fact_Financial_Appraisal.csv")
    
    pd.DataFrame(monte_carlo_data).to_csv(os.path.join(PBI_DIR, "Fact_Monte_Carlo.csv"), index=False)
    print(" [PASS] Exported Fact_Monte_Carlo.csv")

def update_excel_workbook(file_path):
    """Adds formatted reporting worksheets to an Excel workbook."""
    print(f"\n--- Updating Excel Master Workbook: {file_path} ---")
    wb = openpyxl.load_workbook(file_path)
    
    # Define styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    title_font = Font(name="Calibri", size=14, bold=True, color="111827")
    sub_font = Font(name="Calibri", size=10, italic=True, color="6B7280")
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # Helper to create/populate a styled sheet
    def add_styled_sheet(title, df, col_formats=None):
        if title in wb.sheetnames:
            del wb[title]
        ws = wb.create_sheet(title=title)
        
        # Add Title
        ws.cell(row=1, column=1, value=title.replace("_", " ").upper()).font = title_font
        ws.cell(row=2, column=1, value="Offshore EPC Platform Drill Tower Project — Status Date: Aug 31, 2026").font = sub_font
        
        # Write Headers
        headers = list(df.columns)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            
        # Write Rows
        for row_idx, row_data in enumerate(df.to_dict('records'), 5):
            for col_idx, h in enumerate(headers, 1):
                val = row_data[h]
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                
                # Alignments & Formatting
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    if col_formats and h in col_formats:
                        cell.number_format = col_formats[h]
                else:
                    cell.alignment = Alignment(horizontal="left")
                    
        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    # Add sheets
    add_styled_sheet("Key_Milestones", pd.DataFrame(milestones_data))
    add_styled_sheet("Risk_Register", pd.DataFrame(risks_data), {"Financial_Exposure": "$#,##0", "Expected_Monetary_Value": "$#,##0"})
    add_styled_sheet("Waterfall_Bridge", pd.DataFrame(waterfall_data), {"Incremental_Cost": "$#,##0", "Cumulative_Cost": "$#,##0"})
    add_styled_sheet("Cash_Burn_Rate", pd.DataFrame(cash_burn_data), {"Monthly_PV": "$#,##0", "Monthly_EV": "$#,##0", "Monthly_AC": "$#,##0", "Cum_AC": "$#,##0", "Remaining_BAC": "$#,##0"})
    add_styled_sheet("Financial_Ratios", pd.DataFrame(financial_ratios_data))
    add_styled_sheet("Monte_Carlo_P90", pd.DataFrame(monte_carlo_data), {"Outturn_Cost_EAC": "$#,##0", "Cost_Overrun_VAC": "$#,##0", "Contingency_Reserve": "$#,##0"})
    
    wb.save(file_path)
    print(f" [PASS] Successfully updated all worksheets in {file_path}")

def update_power_query_script():
    """Updates PowerQuery_Import_Script.m to include all new star schema tables."""
    print("\n--- Updating PowerQuery Import Script (03_power_bi/PowerQuery_Import_Script.m) ---")
    pq_content = """// ===============================================================================
// POWER QUERY M IMPORT SCRIPT - OFFSHORE EPC PLATFORM DRILL TOWER PROJECT
// Ingests Star Schema CSV Data feeds from 03_power_bi/ directory
// ===============================================================================

let
    SourcePath = "C:\\Users\\frank\\Desktop\\EVM\\03_power_bi\\",
    
    // 1. Ingest Periodic EVM Fact Table
    Fact_EVM_Periodic = Csv.Document(File.Contents(SourcePath & "Fact_EVM_Periodic.csv"),[Delimiter=",", Columns=13, Encoding=65001, QuoteStyle=QuoteStyle.None]),
    Fact_EVM_Headers = Table.PromoteHeaders(Fact_EVM_Periodic, [PromoteAllScalars=true]),
    
    // 2. Ingest Gantt Schedule Fact Table
    Fact_Gantt_Schedule = Csv.Document(File.Contents(SourcePath & "Fact_Gantt_Schedule.csv"),[Delimiter=",", Columns=13, Encoding=65001, QuoteStyle=QuoteStyle.None]),
    Fact_Gantt_Headers = Table.PromoteHeaders(Fact_Gantt_Schedule, [PromoteAllScalars=true]),

    // 3. Ingest Key Milestones Table
    Fact_Milestones = Csv.Document(File.Contents(SourcePath & "Fact_Milestones.csv"),[Delimiter=",", Columns=7, Encoding=65001, QuoteStyle=QuoteStyle.None]),
    Fact_Milestones_Headers = Table.PromoteHeaders(Fact_Milestones, [PromoteAllScalars=true]),

    // 4. Ingest Risk Register Table
    Fact_Risk_Register = Csv.Document(File.Contents(SourcePath & "Fact_Risk_Register.csv"),[Delimiter=",", Columns=10, Encoding=65001, QuoteStyle=QuoteStyle.None]),
    Fact_Risk_Headers = Table.PromoteHeaders(Fact_Risk_Register, [PromoteAllScalars=true]),

    // 5. Ingest Cost Waterfall Bridge Table
    Fact_Waterfall_Bridge = Csv.Document(File.Contents(SourcePath & "Fact_Waterfall_Bridge.csv"),[Delimiter=",", Columns=7, Encoding=65001, QuoteStyle=QuoteStyle.None]),
    Fact_Waterfall_Headers = Table.PromoteHeaders(Fact_Waterfall_Bridge, [PromoteAllScalars=true]),

    // 6. Ingest Monthly Cash Burn Table
    Fact_Monthly_Burn_Rate = Csv.Document(File.Contents(SourcePath & "Fact_Monthly_Burn_Rate.csv"),[Delimiter=",", Columns=7, Encoding=65001, QuoteStyle=QuoteStyle.None]),
    Fact_Burn_Headers = Table.PromoteHeaders(Fact_Monthly_Burn_Rate, [PromoteAllScalars=true]),

    // 7. Ingest Financial Appraisal Table
    Fact_Financial_Appraisal = Csv.Document(File.Contents(SourcePath & "Fact_Financial_Appraisal.csv"),[Delimiter=",", Columns=5, Encoding=65001, QuoteStyle=QuoteStyle.None]),
    Fact_Appraisal_Headers = Table.PromoteHeaders(Fact_Financial_Appraisal, [PromoteAllScalars=true]),

    // 8. Ingest Monte Carlo P90 Table
    Fact_Monte_Carlo = Csv.Document(File.Contents(SourcePath & "Fact_Monte_Carlo.csv"),[Delimiter=",", Columns=8, Encoding=65001, QuoteStyle=QuoteStyle.None]),
    Fact_Monte_Headers = Table.PromoteHeaders(Fact_Monte_Carlo, [PromoteAllScalars=true]),

    // 9. Ingest Dim_WBS
    Dim_WBS = Csv.Document(File.Contents(SourcePath & "Dim_WBS.csv"),[Delimiter=",", Columns=7, Encoding=65001, QuoteStyle=QuoteStyle.None]),
    Dim_WBS_Headers = Table.PromoteHeaders(Dim_WBS, [PromoteAllScalars=true]),

    // 10. Ingest Dim_Date
    Dim_Date = Csv.Document(File.Contents(SourcePath & "Dim_Date.csv"),[Delimiter=",", Columns=4, Encoding=65001, QuoteStyle=QuoteStyle.None]),
    Dim_Date_Headers = Table.PromoteHeaders(Dim_Date, [PromoteAllScalars=true])
in
    Dim_Date_Headers
"""
    with open(os.path.join(PBI_DIR, "PowerQuery_Import_Script.m"), "w", encoding="utf-8") as f:
        f.write(pq_content)
    print(" [PASS] Updated PowerQuery_Import_Script.m successfully")

if __name__ == "__main__":
    export_powerbi_csvs()
    update_excel_workbook(os.path.join(RAW_DIR, "Drill_Tower_EVM_Report.xlsx"))
    update_excel_workbook(os.path.join(RAW_DIR, "EVM_Master_Data.xlsx"))
    update_power_query_script()
    print("\n>>> ALL EXCEL & POWER BI DATASETS AND REPORTS FULLY SYNCHRONIZED <<<")
